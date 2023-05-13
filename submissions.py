'''

This program is free software; you can redistribute it and/or
modify it under the terms of the GNU General Public License
as published by the Free Software Foundation; either version 2
of the License, or (at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program; if not, If not, see <https://www.gnu.org/licenses/>.

'''

import requests
import http.cookiejar
import time
import json
import openpyxl
import decimal
import os


def prepare_submissions():
    session = requests.Session()
    session.cookies = http.cookiejar.MozillaCookieJar('cookies.txt')
    headers = session.headers
    session.cookies.load()
    
    csrftoken = session.cookies._cookies['leetcode.com']['/']['csrftoken'].value
    
    headers['Host'] = 'leetcode.com'
    headers['User-Agent'] = 'Mozilla/5.0 (X11; Linux x86_64; rv:78.0) Gecko/20100101 Firefox/78.0'
    headers['x-csrftoken'] = csrftoken
    
    problems = request(session, 'GET', 'https://leetcode.com/api/problems/all/')
    
    accepted_submissions = []
    
    json_obj = problems.json()
    stat_status_pairs = json_obj['stat_status_pairs']
    for pair in stat_status_pairs:
        if pair['status'] == 'ac':
            accepted_submissions.append({
                'title': pair['stat']['question__title'],
                'slug': pair['stat']['question__title_slug'],
                'difficulty': pair['difficulty']['level'] # 1 - easy, 2 - medium, 3 - hard
            })
    
    return session, accepted_submissions


def actualize(session, accepted_submissions):
    headers = session.headers
    languages = {}
    parsed = []

    i = 1
    for accepted_submission in accepted_submissions:
        leetcode_api_query = {
            "operationName":"Submissions",
            "query":"query Submissions($offset: Int!, $limit: Int!, $lastKey: String, $questionSlug: String!) {  submissionList(offset: $offset, limit: $limit, lastKey: $lastKey, questionSlug: $questionSlug) {submissions {lang, url, runtime}}}",
            "variables": {
                "lastKey": None,
                "limit": 1000,
                "offset": 0,
                "questionSlug": accepted_submission['slug']
            }
        }

        headers['Referer'] = 'https://leetcode.com/problems/' + accepted_submission['slug'] + '/'

        leetcode_api_response = request(session, 'POST', 'https://leetcode.com/graphql', data = json.dumps(leetcode_api_query))

        valid_runtime = lambda submission: submission['runtime'] != 'N/A'

        json_obj = leetcode_api_response.json()
        json_obj['data']['submissionList']['submissions'] = [submission for submission in json_obj['data']['submissionList']['submissions'] if valid_runtime(submission)]

        current_problem_submissions_by_lang = {}
        for submission in json_obj['data']['submissionList']['submissions']:
            submission['runtime'] = int(submission['runtime'].replace(' ms', ''))
            if submission['lang'] not in current_problem_submissions_by_lang or submission['runtime'] < current_problem_submissions_by_lang[submission['lang']]['runtime']:
                current_problem_submissions_by_lang[submission['lang']] = submission
                if submission['lang'] not in languages:
                    newLang = {
                        'lang': submission['lang'],
                        'submissions': []
                    }
                    languages[submission['lang']] = newLang
                    parsed.append(newLang)

        for lang, submission in current_problem_submissions_by_lang.items():
            submission_url = submission['url']

            headers['Referer'] = 'https://leetcode.com/problems/' + accepted_submission['slug'] + '/submissions/'

            submission_detail = request(session, 'GET', 'https://leetcode.com' + submission_url)

            distribution_str = str_between(submission_detail.text, 'runtimeDistributionFormatted: \'', '\',\n')
            if len(distribution_str) > 33:
                json_obj = json.loads(distribution_str.replace('\\u0022', '"'))
                runtimes_distribution = json_obj['distribution'] # [0] - ms, [1] - share
                percent_beats = decimal.Decimal(100)
                for runtime in runtimes_distribution:
                    if submission['runtime'] > int(runtime[0]):
                        percent_beats = percent_beats - decimal.Decimal(str(runtime[1]))
                    else:
                        break

                languages[lang]['submissions'].append({
                    'slug': accepted_submission['slug'],
                    'difficulty': accepted_submission['difficulty'],
                    'problemName': accepted_submission['title'],
                    'beats': percent_beats
                })

        yield '1'
        i = i + 1

    if len(accepted_submissions) > 0:
        wb = openpyxl.Workbook()
        for sheetname in wb.worksheets:
            wb.remove(sheetname)
        
        for language in parsed:
            ws = wb.create_sheet(language['lang'])
            
            for submission in language['submissions']:
                ws.append([submission['slug'], submission['difficulty'], submission['problemName'], submission['beats']])
        
        wb.save('submissions.xlsx')
    
    yield '1'


def cached():
    result = []
    if os.path.exists('submissions.xlsx'):
        wb = openpyxl.load_workbook('submissions.xlsx')
        for ws in wb.worksheets:
            newLang = {
                'lang': ws.title,
                'submissions': []
            }
            result.append(newLang)
            for row in ws.iter_rows(min_row=1, max_col=4, max_row=ws.max_row, values_only=True):
                newLang['submissions'].append({
                    'slug': row[0],
                    'difficulty': row[1],
                    'problemName' :row[2],
                    'beats': decimal.Decimal(row[3])
                })
    
    return result


def str_between(string, str1, str2):
    start_pos = string.find(str1) + len(str1)
    end_pos = string.find(str2, start_pos)
    return string[start_pos:end_pos]


def request(session, method, url, data = None):
    if method.upper() == 'POST':
        session.headers['content-type'] = 'application/json'
        session.headers['Content-Length'] = str(len(str(data)))

    response = session.request(method, url, data = data)

    while (response.status_code == 429):
        print('Rate limit exceeded, slowing down requests')
        time.sleep(61)
        response = session.request(method, url, data = data)

    if response.status_code != 200:
        raise Exception(response.text)

    session.headers.pop('content-type', None)
    session.headers.pop('Content-Length', None)

    return response
